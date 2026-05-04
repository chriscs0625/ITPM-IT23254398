import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

# Initialize workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Test Cases"

# Define columns
headers = [
    "TC ID", 
    "Application Feature Tested", 
    "Input", 
    "Expected output", 
    "Actual output", 
    "Status", 
    "Assumption for Expected Output"
]
ws.append(headers)

# Raw data exactly as requested
csv_data = [
    "Pos_0001,Document conversion,Upload valid DOCX file and click convert,File successfully converts to PDF and download link appears,File converted to PDF and download link appeared successfully,Pass,NA",
    "Neg_0001,Document conversion,Upload an unsupported HTML file,System rejects file and shows Invalid format message,System rejected the file and displayed invalid format message,Pass,NA",
    "Neg_0002,Document conversion,Click convert without uploading any file,System shows error message prompting to upload a file first,Error message appeared prompting file upload,Pass,NA",
    "Pos_0002,PDF editing,Add a text box to the PDF and type Test,Text box appears on the PDF and displays Test,Text box appeared and displayed Test correctly,Pass,NA",
    "Neg_0003,PDF editing,Try to save a PDF without making any changes,System should indicate that no changes were made to save,System indicated no changes were made,Pass,NA",
    "Neg_0004,PDF editing,Place a text box outside the PDF page canvas,Text box is constrained to the PDF page boundaries,Text box was constrained within page boundaries,Pass,NA",
    "Pos_0003,Image resizing,Enter width 500 and height 500,Image preview resizes to the specified 500x500 dimensions,Image resized to 500x500 dimensions successfully,Pass,NA",
    "Pos_0004,Image resizing,Select Maintain aspect ratio and change width,Height automatically updates to maintain the original aspect ratio,Height updated automatically to maintain aspect ratio,Pass,NA",
    "Neg_0005,Image resizing,Enter alphabetical characters abc in the width field,Input is rejected and only numeric values are allowed,Input was rejected and only numeric values accepted,Pass,NA",
    "Neg_0006,Image resizing,Enter a negative value -100 for height,System displays a validation error for negative dimensions,Validation error displayed for negative dimension input,Pass,NA",
    "Pos_0005,Cropping,Draw a valid crop rectangle and click apply,Image is cropped exactly to the defined rectangle,Image was cropped correctly to the defined rectangle,Pass,NA",
    "Pos_0006,Cropping,Select predefined 1:1 format and apply crop,Crop boundaries lock into a perfect square,Crop boundaries locked into a perfect square,Pass,NA",
    "Neg_0007,Cropping,Attempt to drag crop handles completely off the image,Crop handles automatically snap back to the image edges,Crop handles snapped back to image edges automatically,Pass,NA",
    "Neg_0008,Cropping,Click apply crop without drawing any crop box,System warns the user to select an area to crop first,Warning appeared asking user to select a crop area first,Pass,NA",
    "Pos_0007,Compression,Adjust compression slider to 30%,Preview updates and estimated file size is appropriately reduced,Preview updated and file size reduced appropriately,Pass,NA",
    "Pos_0008,Compression,Click reset after compressing,Image preview completely reverts to original uncompressed quality,Image preview reverted to original uncompressed quality,Pass,NA",
    "Neg_0009,Compression,Manually type 200% into compression text box,System caps the value at the maximum 100% allowed limit,System capped the value at 100% maximum limit,Pass,NA",
    "Neg_0010,Compression,Drag the slider down to 0%,System enforces a minimum of 1% compression logic,System enforced minimum 1% compression value,Pass,NA",
    "Pos_0009,Image format conversion,Select PNG as target format for a JPG image,Image converts and downloads successfully as PNG,Image converted and downloaded as PNG successfully,Pass,NA",
    "Neg_0010,Image format conversion,Upload a corrupted image file for conversion,System rejects the file and shows an error message,System rejected the corrupted file with an error message,Pass,NA",
    "Neg_0011,Image format conversion,Attempt conversion without uploading any file,System prompts user to upload a file before converting,System prompted user to upload a file first,Pass,NA",
    "Pos_0010,Meme generation,Upload valid image and add top and bottom text,Meme is generated with text overlaid correctly on image,Meme generated with text overlaid correctly,Pass,NA",
    "Neg_0012,Meme generation,Attempt to generate meme without uploading image,System should prompt user to upload an image first,System prompted user to upload an image,Pass,Image is required before meme generation",
    "Neg_0013,Meme generation,Upload unsupported file type like .pdf for meme,System should reject the file and display format error,System rejected the file and displayed format error,Pass,Meme tool accepts images only",
    "Pos_0011,Color picker,Upload valid image and click on a color,Tool displays the correct hex and RGB values of selected color,Correct hex and RGB values displayed for selected color,Pass,NA",
    "Neg_0014,Color picker,Open color picker without uploading any image,System should prompt user to upload an image first,System prompted user to upload an image first,Pass,Image is required for color picking",
    "Neg_0015,Color picker,Upload unsupported file type like .txt to color picker,System should reject the file and show format error,System rejected the file and showed format error,Pass,Color picker accepts images only",
    "Pos_0025,Image rotation,Upload valid image and rotate,Image should rotate correctly,The uploaded image rotated successfully,Pass,NA",
    "Neg_0026,Image rotation,Click rotate without uploading image,System should prevent rotation,The system blocked rotation and requested image upload,Pass,Image is required before rotation",
    "Neg_0027,Image rotation,Upload unsupported .txt file,System should reject invalid file,The .txt file was rejected by the rotation tool,Pass,Rotation works only for images",
    "Pos_0028,Image flipping,Upload valid image and flip,Image should flip correctly,The image flipped successfully and preview updated,Pass,NA",
    "Neg_0029,Image flipping,Click flip without uploading image,System should prevent flipping,Validation message appeared asking for image upload,Pass,Image is required before flipping",
    "Neg_0030,Image flipping,Upload unsupported file type,System should reject file,Unsupported file was rejected successfully,Pass,Flip tool should accept images only",
    "Pos_0031,Image resizing,Upload JPG and resize with custom dimensions,Resized image should generate correctly,JPG resized successfully using custom dimensions,Pass,NA",
    "Neg_0032,PDF editing,Upload password-protected PDF,System should show error or password request,System displayed restriction message for protected PDF,Pass,Protected PDFs may require password access",
    "Pos_0033,Compression,Upload valid JPG and compress,Image should compress and be downloadable,JPG compressed successfully and download was available,Pass,NA"
]

# Append row by row
for row_string in csv_data:
    row = row_string.split(",")
    ws.append(row)

# Set Column Widths
widths = {'A': 12, 'B': 25, 'C': 35, 'D': 45, 'E': 45, 'F': 10, 'G': 40}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# Styles
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue
header_font = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_align = Alignment(vertical="top", wrap_text=True)

# Apply Styles
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border
        if cell.row == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        else:
            cell.alignment = data_align

# Save the workbook
output_path = Path("Manual Test Cases for Option 2.xlsx").resolve()
wb.save(output_path)
print(f"File created successfully at {output_path} with {len(csv_data)} data rows.")
