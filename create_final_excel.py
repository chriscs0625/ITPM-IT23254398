import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

# Create a new Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Test Cases"

# Define Headers
headers = [
    "TC ID", 
    "Application Feature Tested", 
    "Input", 
    "Expected output", 
    "Actual output", 
    "Status", 
    "Assumption for Expected Output"
]
ws.append(headers)

# Define Borders and Alignments
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Format headers
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border

# Freeze header row
ws.freeze_panes = "A2"

# 36 Test Cases Exactly (10 features * 3 cases + 6 extra cases)
# Only UI/Functionality scenarios, NO backend/security/network/size limits
raw_data = [
    # 1. Document conversion (1 Pos, 2 Neg) = 3
    ("P", "Document conversion", "Upload valid DOCX file", "UI shows success state and provides PDF download link", "TBD", "Not Executed", "NA"),
    ("N", "Document conversion", "Upload unsupported file format (.txt)", "UI displays 'Invalid file format' error message", "TBD", "Not Executed", "NA"),
    ("N", "Document conversion", "Drag and drop multiple files when only one is allowed", "UI warns that only one file is permitted at a time", "TBD", "Not Executed", "NA"),

    # 2. PDF editing (1 Pos, 2 Neg) = 3
    ("P", "PDF editing", "Add text annotation tool to canvas", "Text renders correctly on the PDF preview canvas", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Apply text annotation completely outside PDF canvas bounds", "Text is clipped or restricted to the canvas boundaries", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Click save changes without making any edits", "Save button is disabled or notifies user of no changes", "TBD", "Not Executed", "NA"),

    # 3. Image resizing (2 Pos, 2 Neg) = 4 [1 Extra]
    ("P", "Image resizing", "Specify exactly 800x600 px dimensions", "Image preview correctly scales to the new dimensions", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Enter negative number (-50) in height field", "Input field rejects negative input or shows validation", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Type alphabetic characters ('abc') in width", "Letters are ignored by the numeric input field", "TBD", "Not Executed", "NA"),
    ("P", "Image resizing", "Check 'Maintain Aspect Ratio' and change width", "Height automatically calculates and updates proportionately", "TBD", "Not Executed", "NA"),

    # 4. Cropping (1 Pos, 3 Neg) = 4 [1 Extra]
    ("P", "Cropping", "Draw crop box and click apply crop", "Image bounds visually update to the selected crop area", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Drag crop box partially outside image preview", "Crop box perfectly snaps and restricts to preview edges", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Drag crop handles so width is practically 0", "Handles restricts to a minimum visual crop size (e.g. 1x1)", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Double click crop button without drawing crop box", "Validation message prompts user to draw a box first", "TBD", "Not Executed", "NA"),

    # 5. Compression (2 Pos, 2 Neg) = 4 [1 Extra]
    ("P", "Compression", "Move compression slider to 50% and apply", "Preview visually updates to reflect compressed image", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Enter 150% in manual compression text box", "Value auto-corrects to maximum 100% or shows warning", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Submit compression with slider at 0%", "Value snaps to minimum 1% without throwing console errors", "TBD", "Not Executed", "NA"),
    ("P", "Compression", "Click reset after applying compression", "Preview perfectly reverts to original uncompressed image", "TBD", "Not Executed", "NA"),

    # 6. Image format conversion (1 Pos, 2 Neg) = 3
    ("P", "Image format conversion", "Select Convert PNG to JPG and click convert", "UI provides a JPG conversion success download button", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Attempt to convert JPG to JPG", "Format option is disabled or shows already in format alert", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Leave target format dropdown completely empty", "Validation states 'Please select a target format'", "TBD", "Not Executed", "NA"),

    # 7. Meme generation (1 Pos, 3 Neg) = 4 [1 Extra]
    ("P", "Meme generation", "Enter Top and Bottom text", "Text visibly overlays the image preview with standard meme font", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Submit meme generation without uploading image", "Validation prompts 'Please upload an image first'", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Type text far exceeding typical character length", "Text scales down or wraps instead of overflowing the image", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Submit with both inputs containing only spaces", "Validation flags 'Text cannot be empty'", "TBD", "Not Executed", "NA"),

    # 8. Color picker (1 Pos, 2 Neg) = 3
    ("P", "Color picker", "Click specific specific red pixel on image", "HEX code field updates matching the clicked color", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Click empty background outside the image", "Color picker ignores click and HEX field remains unchanged", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Manually type 'ZZZZZZ' into HEX input field", "Rejects invalid HEX string and reverts to last valid color", "TBD", "Not Executed", "NA"),

    # 9. Image rotation (2 Pos, 2 Neg) = 4 [1 Extra]
    ("P", "Image rotation", "Click 'Rotate Right' button once", "Image visually rotates exactly 90 degrees clockwise", "TBD", "Not Executed", "NA"),
    ("N", "Image rotation", "Input special characters into custom angle field", "Input safely ignores all non-numeric characters", "TBD", "Not Executed", "NA"),
    ("N", "Image rotation", "Interact with rotate controls before uploading image", "Controls dynamically appear greyed out or disabled", "TBD", "Not Executed", "NA"),
    ("P", "Image rotation", "Click 'Rotate Left' exactly 4 times in a row", "Image visually returns identically to its original unrotated state", "TBD", "Not Executed", "NA"),

    # 10. Image flipping (1 Pos, 3 Neg) = 4 [1 Extra]
    ("P", "Image flipping", "Click 'Flip Horizontal' toggle", "Image preview correctly mirrors instantly along the Y-axis", "TBD", "Not Executed", "NA"),
    ("N", "Image flipping", "Click flip buttons continuously with no image", "UI produces no errors and leaves the buttons inactive", "TBD", "Not Executed", "NA"),
    ("N", "Image flipping", "Extremely rapid multi-clicks on Flip Vertical button", "UI queue handles flips sequentially without browser freezing", "TBD", "Not Executed", "NA"),
    ("N", "Image flipping", "Click and drag the flip button instead of pressing", "Button produces standard HTML drag behavior but no flip trigger", "TBD", "Not Executed", "NA"),
]

pos_case_id = 1
neg_case_id = 1

# Append data and apply formatting to every cell
for row_idx, row in enumerate(raw_data, start=2):
    t_type = row[0]
    if t_type == "P":
        tc_id = f"Pos_{pos_case_id:04d}"
        pos_case_id += 1
    else:
        tc_id = f"Neg_{neg_case_id:04d}"
        neg_case_id += 1
    
    row_data = [tc_id] + list(row[1:])
    ws.append(row_data)
    
    # Apply styling to the appended row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx in [1, 6, 7]: # Center TC ID, Status, Assumption
            cell.alignment = center_alignment
        else: # Left align but Center Vertically + Wrap for the rest
            cell.alignment = wrap_alignment

# Format column widths visually
ws.column_dimensions['A'].width = 12  # TC ID
ws.column_dimensions['B'].width = 20  # Application Feature Tested
ws.column_dimensions['C'].width = 35  # Input
ws.column_dimensions['D'].width = 45  # Expected output
ws.column_dimensions['E'].width = 15  # Actual output
ws.column_dimensions['F'].width = 15  # Status
ws.column_dimensions['G'].width = 25  # Assumption

# Final Verification
assert (pos_case_id - 1) + (neg_case_id - 1) == 36, "There must be exactly 36 scenarios!"

# Save as Excel file (.xlsx)
file_path = Path("Manual Test Cases for Option 2.xlsx").resolve()
wb.save(file_path)
print(f"Fixed final Excel file successfully created at: {file_path}")
