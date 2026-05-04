import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

# Create Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Test Cases"

headers = [
    "TC ID", 
    "Application Feature Tested", 
    "Input", 
    "Expected output", 
    "Actual output", 
    "Status", 
    "Assumption for Expected Output"
]
ws.append(headers)

# Styling Definitions
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light green/gray matching professional templates

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Apply header styling
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, size=11)
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = header_fill

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 35 # Taller header

# 36 Test Cases (10 Features, 3 per feature + 6 Extra)
raw_data = [
    # 1. Document conversion
    ("P", "Document conversion", "Upload a valid DOCX file", "System should successfully convert the file to PDF format and provide a download link", "TBD", "Not Executed", "NA"),
    ("N", "Document conversion", "Upload an unsupported file format like .txt", "System should reject the upload and display an 'Invalid file format' error message", "TBD", "Not Executed", "NA"),
    ("N", "Document conversion", "Upload a corrupted DOCX file", "System should handle the error gracefully and display a 'File read error'", "TBD", "Not Executed", "NA"),

    # 2. PDF editing
    ("P", "PDF editing", "Select the text tool and add text to the PDF canvas", "The typed text should render correctly on the PDF preview canvas", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Click save changes without adding any annotations or edits", "The save action should notify the user that no changes were made", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Apply a text annotation completely outside the PDF boundary", "The annotation should be restricted or clipped to the document canvas", "TBD", "Not Executed", "NA"),

    # 3. Image resizing (4 cases)
    ("P", "Image resizing", "Enter valid strict dimensions (e.g., 800 width x 600 height)", "The image preview should correctly scale to the requested dimensions", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Enter negative dimensions (e.g., -500) into the width field", "The input field should reject negative values and display a validation error", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Type alphabetical characters (e.g., 'abc') into the input field", "The numeric field should ignore the keystrokes or show an invalid input error", "TBD", "Not Executed", "NA"),
    ("P", "Image resizing", "Check 'Maintain Aspect Ratio' and double the width value", "The height value should automatically calculate and update proportionally", "TBD", "Not Executed", "NA"),

    # 4. Cropping (4 cases)
    ("P", "Cropping", "Draw a crop box inside the image and click the apply button", "The image visually updates and trims down to the selected crop rectangle", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Attempt to drag the crop box entirely off the image preview", "The crop box automatically snaps to ensure it remains inside image boundaries", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Click the apply crop button without drawing a bounding box", "The system should prompt the user to make a selection first", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Drag the crop corner controls until width and height are zero", "The handles should restrict the crop box to a minimum allowable 1x1 size", "TBD", "Not Executed", "NA"),

    # 5. Compression (4 cases)
    ("P", "Compression", "Move the compression slider to 50% and click apply", "The visual preview renders the compressed image and the file size reduction is shown", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Attempt to type 150% in a manual compression text box", "The value automatically corrects itself to the maximum 100% boundary", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Slide the compression to 0% and execute", "The system snaps the slider to a 1% minimum and prevents zero-value calculation", "TBD", "Not Executed", "NA"),
    ("P", "Compression", "Click the 'Reset' button after compressing the image", "The preview cleanly reverts to the original uncompressed image state", "TBD", "Not Executed", "NA"),

    # 6. Image format conversion (3 cases)
    ("P", "Image format conversion", "Select PNG to JPG conversion and execute", "The UI processes the image and provides a fully valid .jpg file for download", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Try to convert an image into its exact current format (JPG to JPG)", "The UI disables the option or warns the user that no conversion is needed", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Click convert while the target format dropdown is empty", "Validation stops execution and prompts 'Please select a target format'", "TBD", "Not Executed", "NA"),

    # 7. Meme generation (4 cases)
    ("P", "Meme generation", "Enter strings into both Top and Bottom text fields", "Text visibly overlays the image preview with standard meme styling (white, black outline)", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Click generate meme before uploading the background image", "Validation popup correctly prompts 'Please upload an image first'", "TBD", "Not Executed", "Meme generation requires image input"),
    ("N", "Meme generation", "Input text consisting exclusively of blank spaces", "Validation flags the input and states 'Text cannot be entirely empty spaces'", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Enter exceptionally lengthy paragraphs into the top text box", "The text scales down, wraps, or is truncated to fit the image without breaking the layout", "TBD", "Not Executed", "NA"),

    # 8. Color picker (3 cases)
    ("P", "Color picker", "Click a bright red section in the uploaded image", "The color picker successfully identifies and copies the matching HEX/RGB code", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Try to use the color picker on the blank website background", "The tool ignores clicks outside the designated image canvas area", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Manually type an invalid 'XXXXXX' code into the HEX result input", "The input field rejects the invalid entry and maintains the last valid color", "TBD", "Not Executed", "NA"),

    # 9. Image rotation (4 cases)
    ("P", "Image rotation", "Click the 'Rotate Right' 90 degrees button", "The image preview visibly rotates exactly 90 degrees in a clockwise direction", "TBD", "Not Executed", "NA"),
    ("N", "Image rotation", "Input special characters ('@#$') into a custom angle field", "The input field restricts non-numerical keystrokes safely", "TBD", "Not Executed", "NA"),
    ("N", "Image rotation", "Attempt to click rotation controls before uploading an image", "The rotation buttons appear disabled, greyed out, and unclickable", "TBD", "Not Executed", "Image is required before rotation"),
    ("P", "Image rotation", "Click 'Rotate Left' exactly 4 times successively", "The image completes a 360 rotation returning seamlessly to its structural original state", "TBD", "Not Executed", "NA"),

    # 10. Image flipping (4 cases)
    ("P", "Image flipping", "Click the 'Flip Horizontal' action button", "The image preview instantly mirrors along the vertical Y-axis flawlessly", "TBD", "Not Executed", "NA"),
    ("N", "Image flipping", "Continuously click flip buttons when no image is loaded", "The UI produces no execution errors and simply ignores the actions", "TBD", "Not Executed", "Image is required before flipping"),
    ("N", "Image flipping", "Perform extremely rapid multi-clicks on 'Flip Vertical'", "The application natively handles consecutive input queueing without freezing the browser", "TBD", "Not Executed", "NA"),
    ("P", "Image flipping", "Perform a flip action on a tiny 1x1 pixel image", "The system updates successfully handling the bounds without breaking", "TBD", "Not Executed", "NA"),
]

pos_case_id = 1
neg_case_id = 1

# Write data with explicit formatting to prevent the hidden/squished issue
for row_idx, row in enumerate(raw_data, start=2):
    t_type = row[0]
    if t_type == "P":
        tc_id = f"Pos_{pos_case_id:04d}"
        pos_case_id += 1
    else:
        tc_id = f"Neg_{neg_case_id:04d}"
        neg_case_id += 1
    
    row_data = [tc_id] + list(row[1:])
    ws.append(row_data)
    
    # Configure row height to be generous so text does NOT hide
    ws.row_dimensions[row_idx].height = 45 

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        
        if col_idx in [1, 2, 5, 6, 7]: # ID, Feature, Actual Output, Status, Assumption
            cell.alignment = center_align
        else: # Input, Expected Output -> Left align for reading but vertically centered
            cell.alignment = left_align

# Column Widths optimized for visibility
ws.column_dimensions['A'].width = 12  # TC ID
ws.column_dimensions['B'].width = 20  # Application Feature Tested
ws.column_dimensions['C'].width = 38  # Input
ws.column_dimensions['D'].width = 45  # Expected output
ws.column_dimensions['E'].width = 15  # Actual output
ws.column_dimensions['F'].width = 15  # Status
ws.column_dimensions['G'].width = 25  # Assumption

file_path = Path("Manual Test Cases for Option 2.xlsx").resolve()
wb.save(file_path)
print(f"Beautifully formatted Excel file successfully created at: {file_path}")