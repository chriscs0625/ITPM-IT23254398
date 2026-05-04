import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

# Create a new Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sample Template for Option 2"

# Define Headers
headers = [
    "TC ID", 
    "Application Feature Tested", 
    "Input", 
    "Expected output", 
    "Actual output", 
    "Status", 
    "Assumption for Expected Output"
]
ws.append(headers)

# Define Borders and Alignments
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Format headers
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border

# Freeze header row
ws.freeze_panes = "A2"

# 36 Test Cases
raw_data = [
    # 1. Document conversion (1 Pos, 2 Neg)
    ("P", "Document conversion", "Upload valid DOCX file", "System should convert file to PDF format successfully", "TBD", "Not Executed", "Valid file provided"),
    ("N", "Document conversion", "Upload corrupted DOCX file", "System should reject file and show 'Invalid file' error", "TBD", "Not Executed", "NA"),
    ("N", "Document conversion", "Upload DOCX file exceeding 50MB limit", "System should show 'File too large' error message", "TBD", "Not Executed", "Limit is 50MB"),

    # 2. PDF editing (1 Pos, 2 Neg)
    ("P", "PDF editing", "Add text annotation to an uploaded PDF", "Text should appear correctly on the PDF canvas", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Upload password-protected PDF", "System should prompt for password and not crash", "TBD", "Not Executed", "NA"),
    ("N", "PDF editing", "Attempt to save edits on 0-page empty PDF", "System should display error 'Empty document cannot be edited'", "TBD", "Not Executed", "NA"),

    # 3. Image resizing (2 Pos, 3 Neg) -> 2 Extras
    ("P", "Image resizing", "Enter valid dimensions 800x600 px", "Image should resize exactly to 800x600 px rendering clearly", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Enter negative width (e.g., -100) in input", "Input field should reject value or show validation error", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Enter alphabet characters 'abc' in height", "Input field should prevent typing letters", "TBD", "Not Executed", "NA"),
    ("P", "Image resizing", "Select 50% scale from dropdown", "Image should scale down to exactly half its original size", "TBD", "Not Executed", "NA"),
    ("N", "Image resizing", "Enter an excessive scale of 5000%", "System should show validation error 'Scale must be under 1000%'", "TBD", "Not Executed", "NA"),

    # 4. Cropping (2 Pos, 3 Neg) -> 2 Extras
    ("P", "Cropping", "Draw crop box inside image and click apply", "Image should be cropped precisely to the selected region", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Drag crop box partially outside image bounds", "System should properly restrict crop box to image edges", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Set crop dimensions to 0x0 via handles", "Crop action should be disabled or show error message", "TBD", "Not Executed", "NA"),
    ("P", "Cropping", "Select 1:1 aspect ratio constraint and drag", "Crop box should maintain a perfect square shape", "TBD", "Not Executed", "NA"),
    ("N", "Cropping", "Input negative crop coordinates manually", "Warning tooltip should state coordinates must be positive", "TBD", "Not Executed", "NA"),

    # 5. Compression (1 Pos, 3 Neg) -> 1 Extra
    ("P", "Compression", "Upload JPG and select 'High Compression'", "File size metric should drop while visual stays acceptable", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Upload a non-image file (.txt)", "System should reject upload with 'Invalid image format' error", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Select 0% (or empty) quality slider", "Slider should snap to minimum 1% or show validation warning", "TBD", "Not Executed", "NA"),
    ("N", "Compression", "Upload an already highly compressed 1KB image", "System should notify user 'File is already maximally compressed'", "TBD", "Not Executed", "NA"),

    # 6. Image format conversion (1 Pos, 2 Neg)
    ("P", "Image format conversion", "Upload PNG and convert to JPG", "Converted file should download as a valid JPG", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Upload unsupported format (.xyz)", "Error message 'Unsupported file format' should be shown", "TBD", "Not Executed", "NA"),
    ("N", "Image format conversion", "Convert animated GIF to PNG", "User should be warned that animation will be lost", "TBD", "Not Executed", "NA"),

    # 7. Meme generation (1 Pos, 2 Neg)
    ("P", "Meme generation", "Upload valid image and add meme text", "Meme preview should display with text", "TBD", "Not Executed", "NA"),
    ("N", "Meme generation", "Generate meme without uploading image", "System should prevent meme generation", "TBD", "Not Executed", "Meme generation requires image input"),
    ("N", "Meme generation", "Enter very long meme text", "System should handle text properly", "TBD", "Not Executed", "NA"),

    # 8. Color picker (1 Pos, 2 Neg)
    ("P", "Color picker", "Select a color using the picker and copy HEX/RGB value", "Selected color should update correctly and copied value should match the selected color", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Click 'Copy' button without selecting or changing any color", "System should copy the default selected color value correctly or notify the user if no color is selected", "TBD", "Not Executed", "NA"),
    ("N", "Color picker", "Try to manually enter invalid color values OR interact randomly with picker", "System should restrict invalid inputs or maintain valid color range", "TBD", "Not Executed", "NA"),

    # 9. Image rotation (1 Pos, 2 Neg)
    ("P", "Image rotation", "Upload valid image and rotate", "Image should rotate correctly", "TBD", "Not Executed", "NA"),
    ("N", "Image rotation", "Click rotate without uploading image", "System should prevent rotation", "TBD", "Not Executed", "Image is required before rotation"),
    ("N", "Image rotation", "Upload unsupported .txt file", "System should reject invalid file", "TBD", "Not Executed", "Rotation works only for images"),

    # 10. Image flipping (2 Pos, 2 Neg) -> 1 Extra
    ("P", "Image flipping", "Upload valid image and flip", "Image should flip correctly", "TBD", "Not Executed", "NA"),
    ("N", "Image flipping", "Click flip without uploading image", "System should prevent flipping", "TBD", "Not Executed", "Image is required before flipping"),
    ("N", "Image flipping", "Rapid double-click on vertical flip", "System should handle rapid flips natively without freezing", "TBD", "Not Executed", "NA"),
    ("P", "Image flipping", "Click 'Flip Vertical' button on a 1x1 image", "UI should update normally handling extreme bounds correctly", "TBD", "Not Executed", "NA"),
]

pos_case_id = 1
neg_case_id = 1

# Append data and apply formatting to every cell
for row_idx, row in enumerate(raw_data, start=2):
    t_type = row[0]
    if t_type == "P":
        tc_id = f"Pos_{pos_case_id:04d}"
        pos_case_id += 1
    else:
        tc_id = f"Neg_{neg_case_id:04d}"
        neg_case_id += 1
    
    row_data = [tc_id] + list(row[1:])
    ws.append(row_data)
    
    # Apply styling to the appended row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx in [1, 6, 7]: # Center TC ID, Status, Assumption
            cell.alignment = center_alignment
        else: # Left align but Center Vertically + Wrap for the rest
            cell.alignment = wrap_alignment

# Set specific column widths to match the screenshot layout
ws.column_dimensions['A'].width = 12  # TC ID
ws.column_dimensions['B'].width = 16  # Application Feature Tested
ws.column_dimensions['C'].width = 30  # Input
ws.column_dimensions['D'].width = 40  # Expected output
ws.column_dimensions['E'].width = 25  # Actual output
ws.column_dimensions['F'].width = 15  # Status
ws.column_dimensions['G'].width = 25  # Assumption

# Save as Excel file (.xlsx)
file_path = Path("Manual Test Cases for Option 2.xlsx").resolve()
wb.save(file_path)
print(f"Styled Excel file successfully created at: {file_path}")
